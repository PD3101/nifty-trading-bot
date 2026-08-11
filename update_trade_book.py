"""
Update the trade book workbook from console tradebook JSON.

Reads console tradebook JSON (from fetch_tradebook.py), appends new legs to
Raw_Trades, re-matches FIFO per contract in Trade_Log, updates
Open_Positions, refreshes Dashboard/Daily_PnL rows, and saves the workbook
with fullCalcOnLoad so Excel recalculates on open.

Usage:
    python update_trade_book.py [path_to_book.xlsx] [path_to_trades.json]
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

BOOK_PATH = os.getenv(
    'TRADE_BOOK_PATH',
    os.path.expanduser('~/Desktop/Trade_Book_AQ9308.xlsx'))
TRADES_PATH = os.getenv('TRADES_JSON', '/tmp/tb_all.json')

CHARGES_CONFIG = {
    'opt_brokerage_per_order': 20.0,
    'opt_stt_sell_pct': 0.0015,
    'opt_txn_pct': 0.0003553,
    'opt_sebi_per_cr': 10.0,
    'opt_ipft_per_cr': 0.01,
    'opt_stamp_buy_pct': 0.00003,
    'gst_rate': 0.18,
}


def parse_dt(v, fmt):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v:
        try:
            return datetime.strptime(v, fmt)
        except Exception:
            return None
    return None


def normalize(trade):
    """Parse string dates from the console JSON into datetime objects."""
    t = dict(trade)
    t['trade_date'] = parse_dt(t.get('trade_date'), '%Y-%m-%d') or datetime.min
    t['expiry_date'] = parse_dt(t.get('expiry_date'), '%Y-%m-%d') or t['trade_date']
    t['exec_dt'] = parse_dt(t.get('order_execution_time'), '%Y-%m-%dT%H:%M:%S')
    # Precompute per-leg charges (full leg) and charge-per-unit (total / qty),
    # matching the workbook: Trade_Log charges = qty × (buy_cu + sell_cu).
    c = leg_charges(t['quantity'] * t['price'], t['trade_type'])
    t['leg_charges_total'] = c['total']
    t['charge_per_unit'] = c['total'] / t['quantity']
    return t


def leg_charges(turnover, side):
    """Compute charges at full precision (matches the workbook's formulas)."""
    brokerage = CHARGES_CONFIG['opt_brokerage_per_order']
    stt = turnover * CHARGES_CONFIG['opt_stt_sell_pct'] if side == 'sell' else 0.0
    txn = turnover * CHARGES_CONFIG['opt_txn_pct']
    turnover_cr = turnover / 10_000_000.0
    sebi = turnover_cr * CHARGES_CONFIG['opt_sebi_per_cr']
    ipft = turnover_cr * CHARGES_CONFIG['opt_ipft_per_cr']
    stamp = turnover * CHARGES_CONFIG['opt_stamp_buy_pct'] if side == 'buy' else 0.0
    gst = (brokerage + sebi + txn + ipft) * CHARGES_CONFIG['gst_rate']
    total = brokerage + stt + txn + sebi + ipft + stamp + gst
    return {
        'brokerage': brokerage, 'stt': stt, 'txn': txn,
        'sebi': sebi, 'ipft': ipft, 'stamp': stamp,
        'gst': gst, 'total': total,
        'charge_per_unit': round(total / max(1, turnover), 6),
    }


def read_existing_trade_ids(wb):
    ws = wb['Raw_Trades']
    ids = set()
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row and row[0]:
            ids.add(str(row[0]).strip())
    return ids


def console_to_raw_row(t):
    """Convert a normalized console trade to Raw_Trades row values."""
    qty = t['quantity']
    price = t['price']
    turnover = qty * price
    charges = leg_charges(turnover, t['trade_type'])
    return [
        t['trade_id'],            # A: Trade ID
        t['order_id'],            # B: Order ID
        t['tradingsymbol'],       # C: Symbol
        'OPTION',                 # D: Instrument
        t['expiry_date'],         # E: Expiry Date (datetime)
        t['trade_date'],          # F: Trade Date (datetime)
        t['exchange'],            # G: Exchange
        t['segment'],             # H: Segment
        t['trade_type'],          # I: Trade Type
        qty,                      # J: Quantity
        price,                    # K: Price
        turnover,                 # L: Turnover
        charges['brokerage'],     # M: Brokerage
        charges['stt'],           # N: STT
        charges['txn'],           # O: Exchange Txn
        charges['sebi'],          # P: SEBI
        charges['ipft'],          # Q: IPFT
        charges['stamp'],         # R: Stamp
        charges['gst'],           # S: GST
        charges['total'],         # T: Total Charges
        charges['charge_per_unit'],  # U: Charge/Unit
        t['exec_dt'],             # V: Order Execution Time (datetime)
        datetime.now(),           # W: Import Batch Date
    ]


def fifo_match(all_legs):
    legs_by_symbol = defaultdict(list)
    for leg in all_legs:
        legs_by_symbol[leg['tradingsymbol']].append(leg)

    matched = []
    open_positions = []

    for symbol, legs in sorted(legs_by_symbol.items()):
        legs.sort(key=lambda x: (x['trade_date'], x.get('exec_dt') or datetime.min))
        buy_queue = []
        for leg in legs:
            if leg['trade_type'] == 'buy':
                buy_queue.append(leg)
            elif leg['trade_type'] == 'sell':
                remaining = leg['quantity']
                while remaining > 0 and buy_queue:
                    buy = buy_queue[0]
                    if buy['quantity'] > remaining:
                        buy = dict(buy)
                        buy['quantity'] = buy['quantity'] - remaining
                        buy_queue[0] = buy
                        matched_qty = remaining
                    else:
                        matched_qty = buy['quantity']
                        buy_queue.pop(0)
                    matched.append({
                        'symbol': symbol,
                        'quantity': matched_qty,
                        'buy_trade_id': buy['trade_id'],
                        'buy_date': buy['trade_date'],
                        'buy_time': buy.get('exec_dt'),
                        'buy_price': buy['price'],
                        'sell_trade_id': leg['trade_id'],
                        'sell_date': leg['trade_date'],
                        'sell_time': leg.get('exec_dt'),
                        'sell_price': leg['price'],
                        'buy_cu': buy['charge_per_unit'],
                        'sell_cu': leg['charge_per_unit'],
                    })
                    remaining -= matched_qty
        open_positions.extend(buy_queue)

    matched.sort(key=lambda x: (x['sell_date'], x.get('sell_time') or datetime.min))
    return matched, open_positions


def update_workbook():
    trades = [normalize(t) for t in json.load(open(TRADES_PATH))]
    wb = openpyxl.load_workbook(BOOK_PATH)

    existing_ids = read_existing_trade_ids(wb)
    new_ids = [t['trade_id'] for t in trades if t['trade_id'] not in existing_ids]
    print(f'Existing legs: {len(existing_ids)} | New legs: {len(new_ids)}')

    # --- Raw_Trades: rebuild all rows from the console JSON (normalized dates) ---
    ws_rt = wb['Raw_Trades']
    for row in range(4, ws_rt.max_row + 1):
        for col in range(1, ws_rt.max_column + 1):
            ws_rt.cell(row=row, column=col, value=None)
    for i, t in enumerate(sorted(trades, key=lambda x: x['trade_date'])):
        r = i + 4
        for col, val in enumerate(console_to_raw_row(t), 1):
            ws_rt.cell(row=r, column=col, value=val)
    print(f'Raw_Trades data rows: {len(trades)}')

    # --- Trade_Log: full FIFO re-match ---
    matched, open_positions = fifo_match(trades)
    ws_tl = wb['Trade_Log']
    for row in range(4, ws_tl.max_row + 1):
        for col in range(1, ws_tl.max_column + 1):
            ws_tl.cell(row=row, column=col, value=None)

    for i, m in enumerate(matched, 1):
        row = i + 3
        # Charges allocated proportionally per leg (matches workbook formulas):
        # charges = matched_qty × (buy_charge_per_unit + sell_charge_per_unit)
        total_charges = round(m['quantity'] * (m['buy_cu'] + m['sell_cu']), 4)
        gross = round((m['sell_price'] - m['buy_price']) * m['quantity'], 4)
        net = round(gross - total_charges, 4)
        result = 'Win' if net > 0 else ('Loss' if net < 0 else 'Breakeven')
        holding = (m['sell_date'] - m['buy_date']).days
        exit_month = m['sell_date'].strftime('%b-%Y')

        ws_tl.cell(row=row, column=1, value=i)
        ws_tl.cell(row=row, column=2, value=m['symbol'])
        ws_tl.cell(row=row, column=3, value='OPTION')
        ws_tl.cell(row=row, column=4, value=m['sell_date'])
        ws_tl.cell(row=row, column=5, value=m['quantity'])
        ws_tl.cell(row=row, column=6, value=m['buy_trade_id'])
        ws_tl.cell(row=row, column=7, value=m['buy_date'])
        ws_tl.cell(row=row, column=8, value=m['buy_time'])
        ws_tl.cell(row=row, column=9, value=m['buy_price'])
        ws_tl.cell(row=row, column=10, value=m['quantity'] * m['buy_price'])
        ws_tl.cell(row=row, column=11, value=m['sell_trade_id'])
        ws_tl.cell(row=row, column=12, value=m['sell_date'])
        ws_tl.cell(row=row, column=13, value=m['sell_time'])
        ws_tl.cell(row=row, column=14, value=m['sell_price'])
        ws_tl.cell(row=row, column=15, value=m['quantity'] * m['sell_price'])
        ws_tl.cell(row=row, column=16, value=gross)
        ws_tl.cell(row=row, column=17, value=total_charges)
        ws_tl.cell(row=row, column=18, value=net)
        ws_tl.cell(row=row, column=19, value=result)
        ws_tl.cell(row=row, column=20, value=holding)
        ws_tl.cell(row=row, column=21, value=exit_month)

    # --- Open_Positions: clear + write unpaired legs ---
    ws_op = wb['Open_Positions']
    for row in range(4, 100):
        for col in range(1, 14):
            ws_op.cell(row=row, column=col, value=None)
    for i, pos in enumerate(open_positions, 1):
        row = i + 3
        ws_op.cell(row=row, column=1, value=pos['tradingsymbol'])
        ws_op.cell(row=row, column=2, value='OPTION')
        ws_op.cell(row=row, column=3, value=pos['expiry_date'])
        ws_op.cell(row=row, column=4, value='LONG')
        ws_op.cell(row=row, column=5, value=pos['quantity'])
        ws_op.cell(row=row, column=6, value=pos['trade_id'])
        ws_op.cell(row=row, column=7, value=pos['trade_date'])
        ws_op.cell(row=row, column=8, value=pos.get('exec_dt'))
        ws_op.cell(row=row, column=9, value=pos['price'])
        ws_op.cell(row=row, column=10, value=pos['quantity'] * pos['price'])

    # --- Daily_PnL: ensure a row exists for every sell date ---
    ws_d = wb['Daily_PnL']
    sell_dates = sorted({m['sell_date'].date() for m in matched})
    existing_dates = set()
    for r in range(4, ws_d.max_row + 1):
        v = ws_d.cell(row=r, column=1).value
        if isinstance(v, datetime):
            existing_dates.add(v.date())
    for d in sell_dates:
        if d in existing_dates:
            continue
        nr = ws_d.max_row + 1
        ws_d.cell(row=nr, column=1, value=datetime(d.year, d.month, d.day))
        ws_d.cell(row=nr, column=2, value=f'=COUNTIFS(Trade_Log!$L:$L,A{nr},Trade_Log!$K:$K,"<>")')
        ws_d.cell(row=nr, column=3, value=f'=SUMIFS(Trade_Log!$P:$P,Trade_Log!$L:$L,A{nr})')
        ws_d.cell(row=nr, column=4, value=f'=SUMIFS(Trade_Log!$Q:$Q,Trade_Log!$L:$L,A{nr})')
        ws_d.cell(row=nr, column=5, value=f'=SUMIFS(Trade_Log!$R:$R,Trade_Log!$L:$L,A{nr})')
        if nr == 4:
            ws_d.cell(row=nr, column=6, value=f'=E{nr}')
        else:
            ws_d.cell(row=nr, column=6, value=f'=F{nr-1}+E{nr}')
        ws_d.cell(row=nr, column=7, value=f'=IFERROR(B{nr}*A{nr},0)')  # not standard; left simple
        existing_dates.add(d)
    print(f'Daily_PnL rows: {ws_d.max_row - 3}')

    # --- Dashboard: fix hardcoded bits (A2:L2 is one merged cell) ---
    ws_dash = wb['Dashboard']
    ws_dash['A2'] = ('Last updated: 11-Aug-2026      '
                     'Data range: 20-Jul-2026 to 11-Aug-2026')
    ws_dash['A8'] = len(open_positions) if open_positions else 0
    # Charges breakdown on Dashboard is formula-based on Raw_Trades; leave as-is.

    # --- Force Excel to recalculate on open ---
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception as e:
        print('calc warn:', e)

    wb.save(BOOK_PATH)
    print(f'\n✅ Updated & saved: {BOOK_PATH}')
    print(f'   Trade_Log matched: {len(matched)} | Open: {len(open_positions)}')
    for p in open_positions:
        print(f"     {p['tradingsymbol']} LONG qty={p['quantity']} @ {p['price']} (id={p['trade_id']})")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        BOOK_PATH = sys.argv[1]
    if len(sys.argv) > 2:
        TRADES_PATH = sys.argv[2]
    update_workbook()
